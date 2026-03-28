#!/usr/bin/env python3

import argparse
import json
import os
import zlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment


def load_doc_json(path: Path):
    return json.loads(path.read_text())


def decode_lakesheet(obj: dict):
    errors = []
    for field in ("body", "body_draft"):
        body_text = obj.get(field)
        if not body_text:
            continue
        try:
            body = json.loads(body_text) if isinstance(body_text, str) else body_text
            raw = body["sheet"].encode("latin1")
            decoded = zlib.decompress(raw).decode("utf-8")
            return json.loads(decoded)
        except Exception as exc:
            errors.append(f"{field}: {exc}")
    raise ValueError("; ".join(errors) or "missing lakesheet payload")


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = name.translate(str.maketrans({c: "_" for c in '[]:*?/\\'})).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    idx = 1
    while candidate in used:
        suffix = f"_{idx}"
        candidate = (cleaned[: 31 - len(suffix)] + suffix) or f"Sheet{suffix}"
        idx += 1
    used.add(candidate)
    return candidate


def safe_file_name(name: str, used: set[str]) -> str:
    cleaned = name.translate(str.maketrans({c: "_" for c in '<>:"/\\|?*'})).strip() or "untitled"
    candidate = cleaned
    idx = 1
    while candidate in used:
        candidate = f"{cleaned}_{idx}"
        idx += 1
    used.add(candidate)
    return candidate


def find_vault_root(base_output: Path):
    for candidate in [base_output, *base_output.parents]:
        if (candidate / ".meta").is_dir():
            return candidate
    return None


def build_slug_to_markdown(vault_root: Path):
    mapping = {}
    for md in vault_root.rglob("*.md"):
        try:
            for line in md.read_text().splitlines():
                if line.startswith("url: https://www.yuque.com/"):
                    key = line.rsplit("/", 1)[-1].strip()
                    mapping[key] = md
                    break
        except Exception:
            continue
    return mapping


def extract_cell_value(cell):
    value = cell.get("v", "")
    if isinstance(value, dict):
        if value.get("class") == "formula":
            return value.get("value", "")
        return json.dumps(value, ensure_ascii=False)
    return value


def fill_sheet(ws, sheet):
    for row_idx_str, cols in sheet.get("data", {}).items():
        row_idx = int(row_idx_str) + 1
        for col_idx_str, cell in cols.items():
            col_idx = int(col_idx_str) + 1
            excel_cell = ws.cell(row=row_idx, column=col_idx)
            excel_cell.value = extract_cell_value(cell)
            if isinstance(excel_cell.value, str) and "\n" in excel_cell.value:
                excel_cell.alignment = Alignment(wrap_text=True)

    for merge_key, merge in sheet.get("mergeCells", {}).items():
        start_row = int(merge["row"]) + 1
        start_col = int(merge["col"]) + 1
        end_row = start_row + int(merge["rowCount"]) - 1
        end_col = start_col + int(merge["colCount"]) - 1
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def export_one(doc_json_path: Path, output_path: Path):
    obj = load_doc_json(doc_json_path)
    if obj.get("format") != "lakesheet":
        return False

    sheets = decode_lakesheet(obj)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    used_names = set()
    for sheet in sheets:
        ws = wb.create_sheet(title=safe_sheet_name(sheet.get("name", "Sheet"), used_names))
        fill_sheet(ws, sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return True


def rewrite_markdown_as_pointer(md_path: Path, obj: dict, xlsx_path: Path):
    old_text = md_path.read_text() if md_path.exists() else ""
    old_url = ""
    for line in old_text.splitlines():
        if line.startswith("url: "):
            old_url = line.removeprefix("url: ").strip()
            break

    relative_link = os.path.relpath(xlsx_path, start=md_path.parent).replace("\\", "/")
    title = obj.get("title") or obj.get("slug") or md_path.stem
    slug = obj.get("slug") or ""
    content = (
        "---\n"
        f"title: {title}\n"
        f"url: {old_url}\n"
        "export_type: lakesheet\n"
        "---\n\n"
        f"# {title}\n\n"
        "这篇内容在语雀中是电子表格（`lakesheet`），不适合直接导出为 Markdown。\n\n"
        f"- Excel 文件：[{xlsx_path.name}]({relative_link})\n"
        f"- Excel 相对路径：`{relative_link}`\n"
        f"- 原始 slug：`{slug}`\n"
    )
    md_path.write_text(content)


def main():
    parser = argparse.ArgumentParser(description="Export Yuque lakesheet docs to xlsx")
    parser.add_argument("input", help="doc json file or docs directory")
    parser.add_argument("--output-dir", help="output directory for xlsx files")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if input_path.is_file():
        docs = [input_path]
        base_output = Path(args.output_dir).resolve() if args.output_dir else input_path.parent
    else:
        docs = sorted(input_path.glob("*.json"))
        base_output = Path(args.output_dir).resolve() if args.output_dir else input_path.parent.parent.parent.parent / "lakesheet-xlsx"

    vault_root = find_vault_root(base_output)
    slug_to_markdown = build_slug_to_markdown(vault_root) if vault_root else {}

    used_files = set()
    count = 0
    rewritten = 0
    failed = 0
    for doc in docs:
        obj = load_doc_json(doc)
        if obj.get("format") != "lakesheet":
            continue
        title = obj.get("title") or obj.get("slug") or doc.stem
        file_name = safe_file_name(title, used_files)
        out = base_output / f"{file_name}.xlsx"
        try:
            export_one(doc, out)
        except Exception as exc:
            print(f"[WARN] skip {doc}: {exc}")
            failed += 1
            continue
        md_path = slug_to_markdown.get(obj.get("slug", ""))
        if md_path:
            rewrite_markdown_as_pointer(md_path, obj, out)
            rewritten += 1
        print(out)
        count += 1

    print(f"exported {count} lakesheet files")
    print(f"rewrote {rewritten} markdown pointers")
    print(f"skipped {failed} lakesheet files")


if __name__ == "__main__":
    main()
